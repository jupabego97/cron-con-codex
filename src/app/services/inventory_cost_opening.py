"""Import a certified inventory report as the opening cost basis.

The importer intentionally keeps every source row for auditability. Only positive
quantities with an unambiguous product match and a unit cost create an opening
movement and an open cost layer. Negative rows preserve their price as a reference
but do not create inventory or cost.
"""

from __future__ import annotations

import hashlib
import json
import unicodedata
import uuid
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

EXPECTED_HEADERS = {
    "categoria",
    "item",
    "referencia",
    "cantidad",
    "cantidad en remisiones",
    "unidad",
    "estado",
    "costo promedio",
    "total",
}


@dataclass(frozen=True)
class OpeningInventoryRow:
    source_row_number: int
    item_name: str
    reference: str | None
    category: str | None
    unit: str | None
    item_status: str | None
    reported_quantity: Decimal
    reserved_quantity: Decimal
    unit_cost: Decimal | None
    source_total: Decimal | None


@dataclass(frozen=True)
class OpeningInventoryResult:
    import_run_id: uuid.UUID | None
    status: str
    records_read: int
    records_written: int
    layers_created: int
    exception_count: int
    unmatched_count: int
    dry_run: bool


@dataclass(frozen=True)
class _ProductMatch:
    key: int
    alegra_id: str
    name: str
    match_method: str


class InventoryCostOpeningService:
    def __init__(self, *, session: Session) -> None:
        self._session = session

    def import_xlsx(
        self,
        *,
        tenant_id: uuid.UUID,
        file_path: str | Path,
        cutoff_date: date,
        warehouse_alegra_id: str | None = None,
        dry_run: bool = False,
    ) -> OpeningInventoryResult:
        path = Path(file_path)
        rows = read_opening_inventory_xlsx(path)
        source_hash = _file_hash(path)
        warehouse = self._resolve_warehouse(
            tenant_id=tenant_id, warehouse_alegra_id=warehouse_alegra_id
        )
        products = self._load_products(tenant_id=tenant_id)
        classified = [
            self._classify_row(row, products=products, warehouse=warehouse)
            for row in rows
        ]
        exceptions = [row for row in classified if row["classification"] != "opening_layer"]
        unmatched = [
            row
            for row in classified
            if row["classification"] in {"unmatched", "negative_unmatched"}
        ]
        opening_rows = [row for row in classified if row["classification"] == "opening_layer"]

        if dry_run:
            return OpeningInventoryResult(
                import_run_id=None,
                status="dry_run",
                records_read=len(rows),
                records_written=0,
                layers_created=len(opening_rows),
                exception_count=len(exceptions),
                unmatched_count=len(unmatched),
                dry_run=True,
            )

        existing = self._existing_import(
            tenant_id=tenant_id, cutoff_date=cutoff_date, source_hash=source_hash
        )
        if existing is not None and existing["status"] != "failed":
            return OpeningInventoryResult(
                import_run_id=existing["id"],
                status=existing["status"],
                records_read=existing["records_read"],
                records_written=existing["records_written"],
                layers_created=self._count_layers(existing["id"]),
                exception_count=existing["exception_count"],
                unmatched_count=self._count_opening_classification(
                    existing["id"], "unmatched"
                )
                + self._count_opening_classification(existing["id"], "negative_unmatched"),
                dry_run=False,
            )

        conflict = self._successful_cutoff_import(tenant_id=tenant_id, cutoff_date=cutoff_date)
        if conflict is not None:
            raise ValueError(
                "A successful opening cost import already exists for this tenant and cutoff date "
                f"(run={conflict}). Use the same source file for an idempotent retry."
            )

        import_run_id = existing["id"] if existing is not None else uuid.uuid4()
        if existing is None:
            self._session.execute(
                text(
                    """
                    INSERT INTO inventory_cost_import_runs
                      (id, tenant_id, cutoff_date, source_file_name, source_hash, status,
                       records_read, records_written, exception_count)
                    VALUES (:id, :tenant_id, :cutoff_date, :source_file_name, :source_hash,
                            'running', :records_read, 0, :exception_count)
                    """
                ),
                {
                    "id": import_run_id,
                    "tenant_id": tenant_id,
                    "cutoff_date": cutoff_date,
                    "source_file_name": path.name,
                    "source_hash": source_hash,
                    "records_read": len(rows),
                    "exception_count": len(exceptions),
                },
            )
        else:
            self._session.execute(
                text(
                    """
                    UPDATE inventory_cost_import_runs
                    SET status='running', records_read=:records_read,
                        records_written=0, exception_count=:exception_count,
                        finished_at=NULL, error_message=NULL
                    WHERE id=:id
                    """
                ),
                {
                    "id": import_run_id,
                    "records_read": len(rows),
                    "exception_count": len(exceptions),
                },
            )
        self._session.commit()

        try:
            opening_balance_ids: dict[int, uuid.UUID] = {}
            opening_values = []
            for row in classified:
                opening_id = uuid.uuid4()
                opening_balance_ids[row["source_row_number"]] = opening_id
                opening_values.append(
                    {
                        "id": opening_id,
                        "import_run_id": import_run_id,
                        "tenant_id": tenant_id,
                        "cutoff_date": cutoff_date,
                        **row,
                    }
                )
            self._session.execute(
                text(
                    """
                    INSERT INTO inventory_cost_opening_balances
                      (id, import_run_id, tenant_id, cutoff_date, source_row_number,
                       product_key, product_alegra_id, warehouse_key, warehouse_alegra_id,
                       item_name, reference, category, unit, item_status,
                       reported_quantity, reserved_quantity, opening_quantity, unit_cost,
                       calculated_value, source_total, classification, match_method, notes)
                    VALUES
                      (:id, :import_run_id, :tenant_id, :cutoff_date, :source_row_number,
                       :product_key, :product_alegra_id, :warehouse_key, :warehouse_alegra_id,
                       :item_name, :reference, :category, :unit, :item_status,
                       :reported_quantity, :reserved_quantity, :opening_quantity, :unit_cost,
                       :calculated_value, :source_total, :classification, :match_method, :notes)
                    """
                ),
                opening_values,
            )

            movements = []
            layers = []
            for row in opening_rows:
                movement_id = uuid.uuid4()
                source_row = row["source_row_number"]
                quantity = row["opening_quantity"]
                unit_cost = row["unit_cost"]
                movements.append(
                    {
                        "id": movement_id,
                        "tenant_id": tenant_id,
                        "product_key": row["product_key"],
                        "warehouse_key": row["warehouse_key"],
                        "occurred_on": cutoff_date,
                        "source_id": str(import_run_id),
                        "source_line_number": source_row,
                        "quantity_in": quantity,
                        "quantity_out": Decimal("0"),
                        "unit_cost": unit_cost,
                        "total_cost": row["calculated_value"],
                        "metadata": {
                            "opening_balance_id": str(opening_balance_ids[source_row]),
                            "source_row_number": source_row,
                            "source_file_name": path.name,
                        },
                    }
                )
                layers.append(
                    {
                        "id": uuid.uuid4(),
                        "tenant_id": tenant_id,
                        "product_key": row["product_key"],
                        "warehouse_key": row["warehouse_key"],
                        "movement_id": movement_id,
                        "opened_on": cutoff_date,
                        "original_quantity": quantity,
                        "remaining_quantity": quantity,
                        "unit_cost": unit_cost,
                    }
                )
            if movements:
                self._session.execute(
                    text(
                        """
                        INSERT INTO inventory_cost_movements
                          (id, tenant_id, product_key, warehouse_key, occurred_on,
                           movement_type, source_type, source_id, source_line_number,
                           quantity_in, quantity_out, unit_cost, total_cost, cost_method,
                           confidence, metadata)
                        VALUES
                          (:id, :tenant_id, :product_key, :warehouse_key, :occurred_on,
                           'opening_balance', 'inventory_cost_opening', :source_id,
                           :source_line_number, :quantity_in, :quantity_out, :unit_cost,
                           :total_cost, 'moving_average', 'certified', CAST(:metadata AS jsonb))
                        """
                    ),
                    [
                        {**movement, "metadata": json.dumps(movement["metadata"])}
                        for movement in movements
                    ],
                )
                self._session.execute(
                    text(
                        """
                        INSERT INTO inventory_cost_layers
                          (id, tenant_id, product_key, warehouse_key, movement_id, opened_on,
                           original_quantity, remaining_quantity, unit_cost, layer_status)
                        VALUES
                          (:id, :tenant_id, :product_key, :warehouse_key, :movement_id, :opened_on,
                           :original_quantity, :remaining_quantity, :unit_cost, 'open')
                        """
                    ),
                    layers,
                )
            status = "succeeded" if not exceptions else "succeeded_with_exceptions"
            self._session.execute(
                text(
                    """
                    UPDATE inventory_cost_import_runs
                    SET status=:status, records_written=:records_written,
                        finished_at=now()
                    WHERE id=:id
                    """
                ),
                {
                    "id": import_run_id,
                    "status": status,
                    "records_written": len(classified),
                },
            )
            self._session.commit()
            return OpeningInventoryResult(
                import_run_id=import_run_id,
                status=status,
                records_read=len(rows),
                records_written=len(classified),
                layers_created=len(opening_rows),
                exception_count=len(exceptions),
                unmatched_count=len(unmatched),
                dry_run=False,
            )
        except Exception as error:
            self._session.rollback()
            self._session.execute(
                text(
                    """
                    UPDATE inventory_cost_import_runs
                    SET status='failed', error_message=:error_message, finished_at=now()
                    WHERE id=:id
                    """
                ),
                {"id": import_run_id, "error_message": str(error)[:4000]},
            )
            self._session.commit()
            raise

    def _resolve_warehouse(
        self, *, tenant_id: uuid.UUID, warehouse_alegra_id: str | None
    ) -> dict[str, Any]:
        rows = [
            dict(row)
            for row in self._session.execute(
                text(
                    """
                    SELECT key, alegra_id, name
                    FROM dim_warehouse
                    WHERE tenant_id=:tenant_id AND is_deleted=false
                    ORDER BY key
                    """
                ),
                {"tenant_id": tenant_id},
            ).mappings()
        ]
        if warehouse_alegra_id is not None:
            rows = [row for row in rows if str(row["alegra_id"]) == warehouse_alegra_id]
        if len(rows) != 1:
            raise ValueError(
                "The opening report must resolve to exactly one active warehouse; "
                f"found {len(rows)}. Pass --warehouse-alegra-id when necessary."
            )
        return rows[0]

    def _load_products(self, *, tenant_id: uuid.UUID) -> list[dict[str, Any]]:
        return [
            dict(row)
            for row in self._session.execute(
                text(
                    """
                    SELECT key, alegra_id, name, reference
                    FROM dim_product
                    WHERE tenant_id=:tenant_id AND is_deleted=false
                    """
                ),
                {"tenant_id": tenant_id},
            ).mappings()
        ]

    def _classify_row(
        self,
        row: OpeningInventoryRow,
        *,
        products: list[dict[str, Any]],
        warehouse: dict[str, Any],
    ) -> dict[str, Any]:
        match, match_note = _match_product(row, products)
        if match is None:
            classification = "negative_unmatched" if row.reported_quantity < 0 else "unmatched"
            if row.reported_quantity == 0:
                classification = "reference_unmatched"
            return self._opening_values(
                row,
                warehouse=warehouse,
                classification=classification,
                match=None,
                match_note=match_note,
            )
        if row.reported_quantity < 0:
            return self._opening_values(
                row,
                warehouse=warehouse,
                classification="negative_exception",
                match=match,
                match_note="Negative quantity ignored by certified opening policy",
            )
        if row.reported_quantity == 0:
            return self._opening_values(
                row,
                warehouse=warehouse,
                classification="reference_only",
                match=match,
                match_note="Zero quantity does not create an opening layer",
            )
        if row.unit_cost is None:
            return self._opening_values(
                row,
                warehouse=warehouse,
                classification="unpriced_exception",
                match=match,
                match_note="Positive quantity has no unit cost",
            )
        return self._opening_values(
            row,
            warehouse=warehouse,
            classification="opening_layer",
            match=match,
            match_note=None,
        )

    @staticmethod
    def _opening_values(
        row: OpeningInventoryRow,
        *,
        warehouse: dict[str, Any],
        classification: str,
        match: _ProductMatch | None,
        match_note: str | None,
    ) -> dict[str, Any]:
        opening_quantity = (
            row.reported_quantity if classification == "opening_layer" else Decimal("0")
        )
        calculated_value = (
            opening_quantity * row.unit_cost if row.unit_cost is not None else None
        )
        return {
            "source_row_number": row.source_row_number,
            "product_key": match.key if match else None,
            "product_alegra_id": match.alegra_id if match else None,
            "warehouse_key": warehouse["key"],
            "warehouse_alegra_id": str(warehouse["alegra_id"]),
            "item_name": row.item_name,
            "reference": row.reference,
            "category": row.category,
            "unit": row.unit,
            "item_status": row.item_status,
            "reported_quantity": row.reported_quantity,
            "reserved_quantity": row.reserved_quantity,
            "opening_quantity": opening_quantity,
            "unit_cost": row.unit_cost,
            "calculated_value": calculated_value,
            "source_total": row.source_total,
            "classification": classification,
            "match_method": match.match_method if match else None,
            "notes": match_note,
        }

    def _existing_import(
        self, *, tenant_id: uuid.UUID, cutoff_date: date, source_hash: str
    ) -> dict[str, Any] | None:
        row = self._session.execute(
            text(
                """
                SELECT id, status, records_read, records_written, exception_count
                FROM inventory_cost_import_runs
                WHERE tenant_id=:tenant_id AND cutoff_date=:cutoff_date AND source_hash=:source_hash
                ORDER BY started_at DESC LIMIT 1
                """
            ),
            {"tenant_id": tenant_id, "cutoff_date": cutoff_date, "source_hash": source_hash},
        ).mappings().first()
        return dict(row) if row else None

    def _successful_cutoff_import(
        self, *, tenant_id: uuid.UUID, cutoff_date: date
    ) -> uuid.UUID | None:
        value = self._session.execute(
            text(
                """
                SELECT id FROM inventory_cost_import_runs
                WHERE tenant_id=:tenant_id AND cutoff_date=:cutoff_date
                  AND status IN ('succeeded', 'succeeded_with_exceptions')
                ORDER BY finished_at DESC LIMIT 1
                """
            ),
            {"tenant_id": tenant_id, "cutoff_date": cutoff_date},
        ).scalar_one_or_none()
        return value

    def _count_layers(self, import_run_id: uuid.UUID) -> int:
        return int(
            self._session.execute(
                text(
                    """
                    SELECT count(*)
                    FROM inventory_cost_layers l
                    JOIN inventory_cost_movements m ON m.id=l.movement_id
                    WHERE m.source_type='inventory_cost_opening' AND m.source_id=:run_id
                    """
                ),
                {"run_id": str(import_run_id)},
            ).scalar_one()
        )

    def _count_opening_classification(self, import_run_id: uuid.UUID, classification: str) -> int:
        return int(
            self._session.execute(
                text(
                    "SELECT count(*) FROM inventory_cost_opening_balances "
                    "WHERE import_run_id=:run_id AND classification=:classification"
                ),
                {"run_id": import_run_id, "classification": classification},
            ).scalar_one()
        )


def read_opening_inventory_xlsx(path: Path) -> list[OpeningInventoryRow]:
    if not path.is_file():
        raise FileNotFoundError(path)
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        raw_headers = next(values, None)
        if raw_headers is None:
            raise ValueError("The workbook has no header row")
        headers = [_normalize_text(value) for value in raw_headers]
        missing = EXPECTED_HEADERS - set(headers)
        if missing:
            raise ValueError(f"Missing expected columns: {', '.join(sorted(missing))}")
        indexes = {header: index for index, header in enumerate(headers)}
        rows: list[OpeningInventoryRow] = []
        for excel_row_number, values_row in enumerate(values, start=2):
            item_name = _text(_cell(values_row, indexes, "item"))
            if not item_name:
                continue
            rows.append(
                OpeningInventoryRow(
                    source_row_number=excel_row_number,
                    item_name=item_name,
                    reference=_text(_cell(values_row, indexes, "referencia")),
                    category=_text(_cell(values_row, indexes, "categoria")),
                    unit=_text(_cell(values_row, indexes, "unidad")),
                    item_status=_text(_cell(values_row, indexes, "estado")),
                    reported_quantity=_decimal(_cell(values_row, indexes, "cantidad"))
                    or Decimal("0"),
                    reserved_quantity=_decimal(
                        _cell(values_row, indexes, "cantidad en remisiones")
                    )
                    or Decimal("0"),
                    unit_cost=_decimal(_cell(values_row, indexes, "costo promedio")),
                    source_total=_decimal(_cell(values_row, indexes, "total")),
                )
            )
        return rows
    finally:
        workbook.close()


def _match_product(
    row: OpeningInventoryRow, products: list[dict[str, Any]]
) -> tuple[_ProductMatch | None, str | None]:
    by_reference: dict[str, list[dict[str, Any]]] = {}
    by_name: dict[str, list[dict[str, Any]]] = {}
    for product in products:
        reference = _normalize_text(product.get("reference"))
        name = _normalize_text(product.get("name"))
        if reference:
            by_reference.setdefault(reference, []).append(product)
        if name:
            by_name.setdefault(name, []).append(product)
    if row.reference:
        candidates = by_reference.get(_normalize_text(row.reference), [])
        if len(candidates) == 1:
            return _product_match(candidates[0], "reference"), None
        if len(candidates) > 1:
            return None, "Reference matches multiple catalog products"
    candidates = by_name.get(_normalize_text(row.item_name), [])
    if len(candidates) == 1:
        return _product_match(candidates[0], "name"), None
    if len(candidates) > 1:
        return None, "Item name matches multiple catalog products"
    return None, "No catalog product match"


def _product_match(product: dict[str, Any], method: str) -> _ProductMatch:
    return _ProductMatch(
        key=int(product["key"]),
        alegra_id=str(product["alegra_id"]),
        name=str(product["name"]),
        match_method=method,
    )


def _cell(row: tuple[Any, ...], indexes: dict[str, int], name: str) -> Any:
    index = indexes[name]
    return row[index] if index < len(row) else None


def _text(value: Any) -> str | None:
    if value is None:
        return None
    result = str(value).strip()
    return result or None


def _decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError) as error:
        raise ValueError(
            f"Invalid numeric value in opening inventory report: {value!r}"
        ) from error


def _normalize_text(value: Any) -> str:
    text_value = _text(value) or ""
    without_accents = "".join(
        character
        for character in unicodedata.normalize("NFKD", text_value)
        if not unicodedata.combining(character)
    )
    return " ".join(without_accents.casefold().split())


def _file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
